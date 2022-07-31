import io
import os
from konlpy.tag import Kkma
from re import match
from datetime import datetime
import streamlit as st
from PIL import Image
from google.cloud import vision
import openpyxl


#구글의 이미지 인식 서비스를 이용하기 위해선 본인의 GCP_key를 등록해야 한다.
#하단의 GCP_key를 등록하는 법 링크
#https://naramp4.github.io/python/ocr/#%ED%8F%B4%EB%8D%94-%EA%B5%AC%EC%A1%B0

#json형태의 key 경로를 하단에 " "사이에 삽입할 것. 
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r" "


# 영수증 이미지 업로드
@st.cache
def load_image(upload_file):
    img = Image.open(upload_file)
    return img


def upload_credit():
    upload_file = st.file_uploader("영수증 사진을 올리세요.", type=["jpg", "png"])
    if upload_file is not None:
        file_details =\
            {"FileName": upload_file.name, "FileType": upload_file.type}
        st.write(file_details)
        img = load_image(upload_file)
        st.image(img, width=None)
        with open(os.path.join("IMG", upload_file.name), "wb") as f:
            f.write(upload_file.getbuffer())
        st.success("Saved File")

        dir_path = os.getcwd()
        path_IMG = os.path.join(dir_path, 'IMG')
        path_text = os.path.join(path_IMG, str(upload_file.name))
        print(path_text)

        # 한글 분석기 사용
        kkma = Kkma()

        # 영수증 사진으로 부터 재료 텍스트 추출
        """Detects text in the file."""
        client = vision.ImageAnnotatorClient()

        with io.open(path_text, 'rb') as image_file:
            content = image_file.read()

        image = vision.Image(content=content)

        response = client.text_detection(image=image)
        texts = response.text_annotations
        # print(texts)
        list_text = texts[0].description

        ex_sent = kkma.sentences(list_text)
        nouns = []
        for sent in ex_sent:
            for noun in kkma.nouns(sent):
                # 단어 전처리 : 2음절 이상, 수사 제외
                if len(str(noun)) >= 1 and not (match('^[-]?[0-9]', noun)):
                    nouns.append(noun)
        print(nouns)



        # 유통기한 엑셀에 구매일 기록
        wb = openpyxl.load_workbook('./data/db/유통기한.xlsx')
        food_data = wb.active

        ctime = os.path.getctime(path_text)

        date_to_compare = datetime.fromtimestamp(ctime)


        row_len = food_data.max_row

        for i in range(row_len):
            if food_data.cell(row=i + 1, column=1).value in nouns:
                food_data.cell(row=i + 1, column=3).value = date_to_compare
                wb.save('./data/db/유통기한.xlsx')


#유통기한 엑셀에 구매시간 기록
def buy_time(path_text):
    #한글 분석기 사용
    kkma = Kkma()

    #영수증 사진으로 부터 재료 텍스트 추출
    """Detects text in the file."""
    client = vision.ImageAnnotatorClient()

    with io.open(path_text, 'rb') as image_file:
        content = image_file.read()

    image = vision.Image(content=content)

    response = client.text_detection(image=image)
    texts = response.text_annotations
    # print(texts)
    list_text = texts[0].description

    ex_sent = kkma.sentences(list_text)
    print(ex_sent)
    nouns = []
    for sent in ex_sent:
        for noun in kkma.nouns(sent):
            # 단어 전처리 : 2음절 이상, 수사 제외
            if len(str(noun)) >= 1 and not (match('^[-]?[0-9]', noun)):
                nouns.append(noun)


    #유통기한 엑셀에 구매경과일 기록
    wb = openpyxl.load_workbook('./data/db/유통기한.xlsx')
    food_data = wb.active

    ctime = os.path.getctime(path_text)
    now_time = datetime.now()
    date_to_compare = datetime.fromtimestamp(ctime)
    date_diff = now_time - date_to_compare

    row_len = food_data.max_row

    for i in range(row_len):
        if food_data.cell(row=i+1, column=1).value in nouns:
            food_data.cell(row=i+1, column=3).value = date_to_compare
            wb.save('./data/db/유통기한.xlsx')

#남은 유통기한 출력
def show_expiration():
    wb = openpyxl.load_workbook('./data/db/유통기한.xlsx')
    food_data = wb.active
    row_len = food_data.max_row
    print(row_len)
    now_time = datetime.now()
    for i in range(1, row_len):
        if food_data.cell(row=i+1, column=3).value is not None and ((food_data.cell(row=i+1, column=3).value-now_time).days)+food_data.cell(row=i+1, column=2).value <= 3 and ((food_data.cell(row=i+1, column=3).value-now_time).days)+food_data.cell(row=i+1, column=2).value>=0 :
            print('{}의 유통기한이 {}일 남았습니다. 늦지 않게 사용하세요.'.format(food_data.cell(row=i+1, column=1).value, ((food_data.cell(row=i+1, column=3).value-now_time).days)+food_data.cell(row=i+1, column=2).value))
            st.write('{}의 유통기한이 {}일 남았습니다. 늦지 않게 사용하세요.'.format(food_data.cell(row=i+1, column=1).value, ((food_data.cell(row=i+1, column=3).value-now_time).days)+food_data.cell(row=i+1, column=2).value))

#영수증 경로 추적 함수
def move_folder(g_ocr):
    dir_path = os.getcwd()
    path_IMG = os.path.join(dir_path, 'IMG')
    path_text = os.path.join(path_IMG, str(g_ocr))
    print(path_text)
    return path_text


## 유통기한 확인품목 버튼 만들기
def button2():
    if st.button("유통기한 확인"):
        show_expiration()

def button3():
    st.button("레시피 검색")


